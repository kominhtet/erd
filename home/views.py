from django.shortcuts import render, get_object_or_404, redirect
from .models import Party, District, Khayine, Township, Constituency, Numberofvoters1, Valid_vote
from django.core.files.storage import FileSystemStorage
from django.core.files.storage import default_storage
import pandas as pd
from django.conf import settings
import os
import openpyxl
from rest_framework import viewsets
from .serializers import ValidVoteSerializer,ConstituencySerializer,DistrictSerializer
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ValidVoteFilter,ConstituencyFilter,DistrictFilter
from django.core.paginator import Paginator
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count,Q

class ConstituencyViewSet(viewsets.ModelViewSet):
    queryset = Constituency.objects.all()
    serializer_class = ConstituencySerializer 
    filter_backends = [DjangoFilterBackend]  
    filterset_class = ConstituencyFilter 
     
class ValidVoteViewSet(viewsets.ModelViewSet):
    queryset = Valid_vote.objects.all()
    serializer_class = ValidVoteSerializer
    filter_backends = [DjangoFilterBackend]  
    filterset_class = ValidVoteFilter 

class DistrictViewSet(viewsets.ModelViewSet):
    queryset = District.objects.all()
    serializer_class = DistrictSerializer
    filter_backends = [DjangoFilterBackend]  
    filterset_class = DistrictFilter 

def party_result(request):
    party = request.GET.get('party', '')
    election_year = request.GET.get('election_year', '')
    
    votes = Valid_vote.objects.all()

    # Filter by party and election year if provided
    if party:
        votes = votes.filter(party__icontains=party)
    if election_year:
        votes = votes.filter(election_year=election_year)

    # Filter winning and losing votes
    winning_votes = votes.filter(win_lose=True)
    losing_votes = votes.filter(win_lose=False)

    # Count total votes, winning votes, and losing votes
    total_votes_count = votes.count()
    winning_votes_count = winning_votes.count()
    losing_votes_count = losing_votes.count()

    votes_per_year = (
        votes.values('election_year')
        .annotate(
            total_votes_count=Count('id'),
            winning_votes_count=Count('id', filter=Q(win_lose=True)),
            losing_votes_count=Count('id', filter=Q(win_lose=False)),
        )
        .order_by('election_year')  # Sort by election year
    )
    context = {
        'votes_per_year': votes_per_year,
        'votes': votes,  # All filtered votes
        'winning_votes': winning_votes,  # Winning votes queryset
        'losing_votes': losing_votes,  # Losing votes queryset
        'show_tables': bool(request.GET),  # Boolean to show tables
        'total_votes_count': total_votes_count,  # Total count
        'winning_votes_count': winning_votes_count,  # Winning count
        'losing_votes_count': losing_votes_count,  # Losing count
    }
    return render(request, 'home/party_result.html', context)
   
def index(request):  
    return render(request, 'home/index.html')

def admin_page(request):  
    return render(request, 'index.html')

def party_activity(request):   
    return render(request, 'home/party_activity.html')

def home(request):
    return render(request, 'home/home.html')

# Party Views
def party_list(request):
    # Total count of parties in Party model
    total_parties = Party.objects.count()
    
    # Party count in Valid_vote per year
    votes_per_year = (
        Valid_vote.objects.values('election_year')  # Group by election year
        .annotate(party_count=Count('party', distinct=True))  # Count distinct parties per year
        .order_by('election_year')  # Optional: order by year
    )
    
    parties = Party.objects.all()
    votes = Valid_vote.objects.all()
    parties_2010 = Valid_vote.objects.filter(election_year=2010).values_list('party', flat=True).distinct()
    parties_2015 = Valid_vote.objects.filter(election_year=2015).values_list('party', flat=True).distinct()
    parties_2020 = Valid_vote.objects.filter(election_year=2020).values_list('party', flat=True).distinct()
    
    context = {
        'parties': parties,
        'votes': votes,  # All filtered votes
        'total_parties': total_parties,  # Total party count in Party model
        'votes_per_year': votes_per_year,
        'parties_2010': parties_2010, 
        'parties_2015': parties_2015, 
        'parties_2020': parties_2020, 
    }
    return render(request, 'home/party_list.html', context)

def party_list2(request):
    total_parties = Party.objects.count()
    parties = Party.objects.all()
    votes = Valid_vote.objects.all()

    # Count distinct parties by year
    party_count_2010 = Valid_vote.objects.filter(election_year=2010).values('party').distinct().count()
    party_count_2015 = Valid_vote.objects.filter(election_year=2015).values('party').distinct().count()
    party_count_2020 = Valid_vote.objects.filter(election_year=2020).values('party').distinct().count()

    # Get distinct party names for the 2010 election
    parties_2010 = Valid_vote.objects.filter(election_year=2010).values_list('party', flat=True).distinct()
    parties_2015 = Valid_vote.objects.filter(election_year=2015).values_list('party', flat=True).distinct()
    parties_2020 = Valid_vote.objects.filter(election_year=2020).values_list('party', flat=True).distinct()
    
    votes_per_year = (
        Valid_vote.objects.values('election_year')  # Group by election year
        .annotate(party_count=Count('party', distinct=True))  # Count distinct parties per year
        .order_by('election_year')  # Optional: order by year
    )

    context = {
        'votes': votes,
        'parties': parties,
        'votes_per_year': votes_per_year,
        'total_parties': total_parties,
        'party_count_2010': party_count_2010,
        'party_count_2015': party_count_2015,
        'party_count_2020': party_count_2020,
        'parties_2010': parties_2010, 
        'parties_2015': parties_2015, 
        'parties_2020': parties_2020, 
    }
    return render(request, 'party_list2.html', context)



def party_detail(request, pk):
    party = get_object_or_404(Party, pk=pk)
    return render(request, 'home/party_detail.html', {'party': party})

def party_create(request):
    if request.method == 'POST':
        # Handle form submission for single party creation
        if 'name' in request.POST:
            name = request.POST['name']
            logo = request.FILES.get('logo')
            Party.objects.create(name=name, logo=logo)
            return redirect('party_list')

        # Handle Excel file upload for bulk party creation
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            filepath = fs.path(filename)

            # Open the Excel file
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Loop through the rows in the Excel file and create Party objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                party_name = row[0]  # Assuming the first column contains the party name
                Party.objects.create(name=party_name)

            return redirect('party_list')

    return render(request, 'home/party_create.html')

def party_update(request, pk):
    party = get_object_or_404(Party, pk=pk)
    if request.method == 'POST':
        party.name = request.POST['name']
        if 'logo' in request.FILES:
            party.logo = request.FILES['logo']
        party.save()
        return redirect('party_detail', pk=party.pk)
    return render(request, 'home/party_update.html', {'party': party})

def party_delete(request, pk):
    party = get_object_or_404(Party, pk=pk)
    party.delete()
    return redirect('party_list')

# Repeat the similar pattern for District, Khayine, Township, and Constituency
def district_list(request):
    districts = District.objects.all()
    return render(request, 'district/district_list.html', {'districts': districts})

def district_create(request):
    if request.method == 'POST':
        # Handle form submission for single district creation
        if 'name' in request.POST and not request.FILES:
            name = request.POST['name']
            District.objects.create(name=name)
            return redirect('district_list')

        # Handle Excel file upload for bulk district creation
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            filepath = fs.path(filename)

            # Open the Excel file
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Loop through the rows in the Excel file and create District objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                district_name = row[0]  # Assuming the first column contains the district name
                District.objects.create(name=district_name)

            # Delete the file after processing
            fs.delete(filepath)

            return redirect('district_list')

    return render(request, 'district/district_create.html')

def district_update(request, pk):
    district = get_object_or_404(District, pk=pk)
    if request.method == 'POST':
        district.name = request.POST['name']
        district.save()
        return redirect('district_list')
    return render(request, 'district/district_update.html', {'district': district})

def district_delete(request, pk):
    district = get_object_or_404(District, pk=pk)
    district.delete()
    return redirect('district_list')

# Khayine List View
def khayine_list(request):
    search_name = request.GET.get('name', '')
    search_district = request.GET.get('district', '')

    khayines = Khayine.objects.all()
    if search_name:
        khayines = khayines.filter(name__icontains=search_name)
    if search_district:
        khayines = khayines.filter(district__name__icontains=search_district)

    paginator = Paginator(khayines, 10)  
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number)

    context = {
        # 'khayines': page_obj.object_list,  # Pass the list of Khayines
        'page_obj': page_obj,             # For pagination controls
    }
    return render(request, 'khayine/khayine_list.html', context)

def constituency_list(request):
    search_name = request.GET.get('name', '')
    search_district = request.GET.get('district', '')
    
    constituencys = Constituency.objects.all()
    if search_name:
        constituencys = constituencys.filter(name__icontains=search_name)
    if search_district:
        constituencys = constituencys.filter(district__name__icontains=search_district)
    
    # Implement pagination
    paginator = Paginator(constituencys, 10)  
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number)  
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'constituency/constituency_list.html', context)

def constituency_update(request, pk):
    constituency = get_object_or_404(Constituency, pk=pk)
    districts = District.objects.all()
    khayines = Khayine.objects.all()
    townships = Township.objects.all()
    
    if request.method == 'POST':
        constituency.name = request.POST['name']
        district_id = request.POST.get('district')
        constituency.district = District.objects.get(id=district_id) if district_id else None
        
        khayine_id = request.POST.get('khayine')
        constituency.khayine = Khayine.objects.get(id=khayine_id) if khayine_id else None
        
        township_id = request.POST.get('township')
        constituency.township = Township.objects.get(id=township_id) if township_id else None
        
        # Save files if provided
        if 'file' in request.FILES:
            constituency.file = request.FILES['file']
        
        if 'detail_file' in request.FILES:
            constituency.detail_file = request.FILES['detail_file']
        
        constituency.save()
        return redirect('constituency_list')
    
    return render(request, 'constituency/constituency_update.html', {
        'constituency': constituency,
        'districts': districts,
        'khayines': khayines,
        'townships': townships,
    })

def constituency_delete(request, pk):
    constituency = get_object_or_404(Constituency, pk=pk)
    constituency.delete()
    return redirect('constituency_list')

def constituency_create(request):
    districts = District.objects.all()
    khayines = Khayine.objects.all()
    townships = Township.objects.all()
    
    if request.method == 'POST':
        # Handle form submission for single district creation
        if 'name' in request.POST:
            name = request.POST['name']
            district_id = request.POST.get('district')
            district = District.objects.get(id=district_id) if district_id else None
            
            khayine_id = request.POST['khayine']
            khayine = Khayine.objects.get(id=khayine_id)
            
            township_id = request.POST['township']
            township = Township.objects.get(id=township_id)
            
            file = request.FILES.get('file')
            detail_file = request.FILES.get('detail_file')
            Constituency.objects.create(name=name, district=district, khayine=khayine, township=township, file=file, detail_file=detail_file )
            return redirect('constituency_list')
    # Handle bulk creation via Excel upload
    if 'file' in request.FILES:
        excel_file = request.FILES['file']

        # Save the uploaded file
        file_path = default_storage.save(excel_file.name, excel_file)

        # Get the absolute path for reading the file
        file_full_path = os.path.join(settings.MEDIA_ROOT, file_path)

        # Load the Excel file using pandas
        try:
            df = pd.read_excel(file_full_path)

            # Iterate through each row in the Excel file
            for index, row in df.iterrows():
                name = row['name'].strip() if 'name' in row and pd.notna(row['name']) else None
                district_id = row['district_id'] if 'district_id' in row and pd.notna(row['district_id']) else None

                if name and district_id:
                    try:
                        # Fetch the District object
                        district = District.objects.get(id=int(district_id))
                        
                        # Create the Constituency object
                        Constituency.objects.create(name=name, district=district)
                    except District.DoesNotExist:
                        print(f"District with ID '{district_id}' does not exist. Skipping row {index + 1}.")
                else:
                    print(f"Invalid data in row {index + 1}: name='{name}', district_id='{district_id}'. Skipping.")

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        # Delete the file after processing
        default_storage.delete(file_path)
        return redirect('constituency_list')
    return render(request, 'constituency/constituency_create.html', {'districts': districts, 'khayines': khayines, 'townships': townships})

# Khayine Create View
def khayine_create(request):
    districts = District.objects.all()

    if request.method == 'POST':
        if 'name' in request.POST and not request.FILES:
            name = request.POST['name']
            district_id = request.POST.get('district')
            district = District.objects.get(id=district_id) if district_id else None
            Khayine.objects.create(name=name, district=district)
            return redirect('khayine_list')

        # Handle Excel file upload for bulk Khayine creation
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            filepath = fs.path(filename)

            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                khayine_name = row[0] 
                district_id = row[1]   

                district = District.objects.get(id=district_id) if district_id else None
                Khayine.objects.create(name=khayine_name, district=district)

            fs.delete(filepath)

            return redirect('khayine_list')

    return render(request, 'khayine/khayine_create.html', {'districts': districts})

# Khayine Update View
def khayine_update(request, pk):
    khayine = get_object_or_404(Khayine, pk=pk)
    districts = District.objects.all()
    if request.method == 'POST':
        khayine.name = request.POST['name']
        district_id = request.POST.get('district')
        khayine.district = District.objects.get(id=district_id) if district_id else None
        khayine.save()
        return redirect('khayine_list')
    return render(request, 'khayine/khayine_update.html', {'khayine': khayine, 'districts': districts})

# Khayine Delete View
def khayine_delete(request, pk):
    khayine = get_object_or_404(Khayine, pk=pk)
    khayine.delete()
    return redirect('khayine_list')

def township_list(request):   
    search_name = request.GET.get('name', '')
    search_khayine = request.GET.get('khayine', '')

    townships = Township.objects.all()
    if search_name:
        townships = townships.filter(name__icontains=search_name)
    if search_khayine:
        townships = townships.filter(khayine__name__icontains=search_khayine)

    paginator = Paginator(townships, 10)  
    page_number = request.GET.get('page')  
    page_obj = paginator.get_page(page_number)

    context = {
        'townships': page_obj.object_list,  # Pass the list of Khayines
        'page_obj': page_obj,             # For pagination controls
    }
    
    return render(request, 'township/township_list.html', context)

def township_create(request):
    khayines = Khayine.objects.all()  # Fetch khayines for the dropdown

    if request.method == 'POST':
        # Handle single township creation
        if 'name' in request.POST and 'khayine' in request.POST:
            name = request.POST['name']
            khayine_id = request.POST['khayine']
            khayine = Khayine.objects.get(id=khayine_id)
            Township.objects.create(name=name, khayine=khayine)
            return redirect('township_list')

        # Handle bulk creation via Excel upload
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            
            # Save the uploaded file
            file_path = default_storage.save(excel_file.name, excel_file)
            
            # Get the absolute path for reading the file
            file_full_path = os.path.join(settings.MEDIA_ROOT, file_path)
            
            # Load the Excel file using pandas
            df = pd.read_excel(file_full_path)

            # Assuming the Excel file has two columns: 'name' and 'khayine_id'
            for index, row in df.iterrows():
                name = row['name']
                khayine_id = row['khayine_id']
                khayine = Khayine.objects.get(id=khayine_id)
                Township.objects.create(name=name, khayine=khayine)

            # Delete the file after processing
            default_storage.delete(file_path)
            return redirect('township_list')

    return render(request, 'township/township_create.html', {'khayines': khayines})

def township_update(request, pk):
    township = get_object_or_404(Township, pk=pk)
    
    if request.method == 'POST':
        township.name = request.POST['name']
        khayine_id = request.POST['khayine']
        township.khayine = Khayine.objects.get(id=khayine_id)
        township.save()
        return redirect('township_list')
    
    khayines = Khayine.objects.all()
    return render(request, 'township/township_update.html', {'township': township, 'khayines': khayines})

def township_delete(request, pk):
    township = get_object_or_404(Township, pk=pk)
    township.delete()
    return redirect('township_list')

def number_of_voters_list1(request):
    voters = Numberofvoters1.objects.all()
    return render(request, 'number_of_voters1/number_of_voters_list.html', {'voters': voters})

def number_of_voters_create1(request):
    if request.method == 'POST':
        # Handle single record creation
        if 'can_vote' in request.POST and not request.FILES:
            can_vote = request.POST.get('can_vote')
            station_vote = request.POST.get('station_vote')
            advance_vote = request.POST.get('advance_vote')
            total_vote = request.POST.get('total_vote')
            constituency_name = request.POST.get('constituency')  # Directly use constituency name
            district_name = request.POST.get('district')          # Directly use district name
            election_year = request.POST.get('election_year')

            # Create the Numberofvoters1 entry directly without querying Constituency and District models
            Numberofvoters1.objects.create(
                can_vote=can_vote,
                station_vote=station_vote,
                advance_vote=advance_vote,
                total_vote=total_vote,
                constituency=constituency_name,  # Directly use name
                district=district_name,          # Directly use name
                election_year=election_year
            )
            return redirect('number_of_voters_list1')

        # Handle Excel file upload for bulk creation
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            filepath = fs.path(filename)

            # Open the Excel file
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Loop through the rows in the Excel file and create Numberofvoters1 objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                can_vote = row[0] if row[0] is not None else 0
                station_vote = row[1] if row[1] is not None else 0
                advance_vote = row[2] if row[2] is not None else 0
                total_vote = row[3] if row[3] is not None else 0
                constituency_name = row[4] if row[4] is not None else ''
                district_name = row[5] if row[5] is not None else ''
                election_year = row[6] if row[6] is not None else None

                # Directly create Numberofvoters1 objects with names from Excel
                Numberofvoters1.objects.create(
                    can_vote=can_vote,
                    station_vote=station_vote,
                    advance_vote=advance_vote,
                    total_vote=total_vote,
                    constituency=constituency_name,  # Directly use name
                    district=district_name,          # Directly use name
                    election_year=election_year
                )

            # Delete the file after processing
            fs.delete(filepath)

            return redirect('number_of_voters_list1')

    return render(request, 'number_of_voters1/number_of_voters_create.html')

def number_of_voters_update1(request, id):
    voter = get_object_or_404(Numberofvoters1, id=id)

    if request.method == 'POST':
        voter.can_vote = request.POST.get('can_vote', voter.can_vote)
        voter.station_vote = request.POST.get('station_vote', voter.station_vote)
        voter.advance_vote = request.POST.get('advance_vote', voter.advance_vote)
        voter.total_vote = request.POST.get('total_vote', voter.total_vote)
        voter.constituency = request.POST.get('constituency', voter.constituency)
        voter.district = request.POST.get('district', voter.district)
        voter.election_year = request.POST.get('election_year', voter.election_year)

        voter.save()
        return redirect('number_of_voters_list1')

    return render(request, 'number_of_voters1/number_of_voters_update.html', {'voter': voter})

def number_of_voters_delete1(request, id):
    voter = get_object_or_404(Numberofvoters1, id=id)
    
    if request.method == 'POST':
        voter.delete()
        return redirect('number_of_voters_list1')

    return render(request, 'number_of_voters1/number_of_voters_delete.html', {'voter': voter})

# List view for valid votes
def valid_vote_list(request):
    votes = Valid_vote.objects.all()
    return render(request, 'valid_vote/valid_vote_list.html', {'votes': votes})

# Create view for valid votes (single or bulk via Excel)
def valid_vote_create(request):
    districts = District.objects.all()
    if request.method == 'POST':
        # Handle single record creation
        if 'station_vote' in request.POST and not request.FILES:
            station_vote = request.POST.get('station_vote')
            advance_vote = request.POST.get('advance_vote')
            total_vote = request.POST.get('total_vote')
            percentage = request.POST.get('percentage')
            candidate = request.POST.get('candidate')
            party = request.POST.get('party')
            constituency = request.POST.get('constituency')
            
            district_id = request.POST.get('district')
            district = District.objects.get(id=district_id) if district_id else None
            
            election_year = request.POST.get('election_year')
            hlutdaw = request.POST.get('hlutdaw')

            # Create the Valid_vote entry
            Valid_vote.objects.create(
                station_vote=station_vote,
                advance_vote=advance_vote,
                total_vote=total_vote,
                percentage=percentage,
                candidate=candidate,
                party=party,
                constituency=constituency,
                district=district,
                election_year=election_year,
                hlutdaw=hlutdaw
            )
            return redirect('valid_vote_list')

        # Handle Excel file upload for bulk creation
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            filepath = fs.path(filename)

            # Open the Excel file
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Loop through rows in Excel and create Valid_vote objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                station_vote = row[0] if row[0] is not None else 0
                advance_vote = row[1] if row[1] is not None else 0
                total_vote = row[2] if row[2] is not None else 0
                percentage = row[3] if row[3] is not None else 0
                candidate = row[4] if row[4] is not None else ''
                party = row[5] if row[5] is not None else ''
                constituency_name = row[6].strip() if row[6] is not None else ''  # Strip whitespace
                district_id = row[7] if row[7] is not None else 0
                election_year = row[8] if row[8] is not None else None
                hlutdaw = row[9] if row[9] is not None else None
                win_lose = bool(row[10]) if row[10] is not None else False
                
                # Fetch the district object
                district = District.objects.get(id=district_id)

                try:
                    # Fetch the constituency object based on its name
                    constituency = Constituency.objects.get(name__iexact=constituency_name)  # Case-insensitive match
                except Constituency.DoesNotExist:
                    print(f"Constituency '{constituency_name}' not found. Skipping this row.")
                    continue

                # Create a Valid_vote object
                Valid_vote.objects.create(
                    station_vote=station_vote,
                    advance_vote=advance_vote,
                    total_vote=total_vote,
                    percentage=percentage,
                    candidate=candidate,
                    party=party,
                    constituency=constituency,
                    district=district,
                    election_year=election_year,
                    hlutdaw=hlutdaw,
                    win_lose=win_lose
                )

            # Delete the file after processing
            fs.delete(filepath)

            return redirect('valid_vote_list')

    return render(request, 'valid_vote/valid_vote_create.html',{'districts': districts})

# Update view for valid vote
def valid_vote_update(request, id):
    vote = get_object_or_404(Valid_vote, id=id)

    if request.method == 'POST':
        vote.station_vote = request.POST.get('station_vote', vote.station_vote)
        vote.advance_vote = request.POST.get('advance_vote', vote.advance_vote)
        vote.total_vote = request.POST.get('total_vote', vote.total_vote)
        vote.percentage = request.POST.get('percentage', vote.percentage)
        vote.candidate = request.POST.get('candidate', vote.candidate)
        vote.party = request.POST.get('party', vote.party)
        vote.constituency = request.POST.get('constituency', vote.constituency)
        vote.district = request.POST.get('district', vote.district)
        vote.election_year = request.POST.get('election_year', vote.election_year)

        vote.save()
        return redirect('valid_vote_list')

    return render(request, 'valid_vote/valid_vote_update.html', {'vote': vote})

# Delete view for valid vote
def valid_vote_delete(request, id):
    vote = get_object_or_404(Valid_vote, id=id)
    
    if request.method == 'POST':
        vote.delete()
        return redirect('valid_vote_list')

    return render(request, 'valid_vote/valid_vote_delete.html', {'vote': vote})
